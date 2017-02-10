import openpyxl

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.get_highest_row):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment !='paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

print(unpaidMembers)

sheet.cell(row=1, column=3).value = 'hello'
